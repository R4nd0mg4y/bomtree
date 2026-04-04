import os, sqlite3, json, io, tempfile
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

st.set_page_config(page_title="BOM Tree", layout="wide", page_icon="⚙️")

DB = os.path.join(os.path.dirname(__file__), "bom.db")

# ── DB ────────────────────────────────────────────────────────────────────────
@st.cache_resource
def get_connection():
    conn = sqlite3.connect(DB, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def get_db():
    return get_connection()

def init_db():
    db = get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS imports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT, col_map TEXT,
        imported_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS parts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        import_id INTEGER REFERENCES imports(id),
        row_no INTEGER, status TEXT, amat_parent TEXT,
        amat_code TEXT NOT NULL, lv INTEGER,
        amat_rev TEXT, fv_code TEXT, fv_rev TEXT, part_name TEXT,
        UNIQUE(import_id, row_no)
    );
    CREATE INDEX IF NOT EXISTS idx_amat ON parts(amat_code);
    CREATE INDEX IF NOT EXISTS idx_parent ON parts(amat_parent);
    """)
    try:
        db.execute("ALTER TABLE imports ADD COLUMN col_map TEXT")
        db.commit()
    except:
        pass

init_db()

DEFAULT_COL_MAP = {
    'status': 1, 'row_no': 2, 'amat_parent': 3, 'lv': 4,
    'amat_code': 5, 'amat_rev': 6, 'fv_code': 7, 'fv_rev': 8, 'part_name': 9,
}
FIELD_LABELS = {
    'status': 'Status', 'row_no': 'Row No.', 'amat_parent': 'AMAT Parent',
    'lv': 'Level (LV)', 'amat_code': 'AMAT Code ★', 'amat_rev': 'AMAT Rev',
    'fv_code': 'FV Code', 'fv_rev': 'FV Rev', 'part_name': 'Part Name',
}

# ── Import ────────────────────────────────────────────────────────────────────
def parse_df(path, sheet=None):
    if path.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(path, header=None, dtype=str, sheet_name=sheet)
    else:
        df = pd.read_csv(path, header=None, dtype=str)
    return df.dropna(how='all').reset_index(drop=True)

def read_by_positions(df, col_map):
    ncols = len(df.columns)
    def g(row, field, default=''):
        pos = col_map.get(field, 0)
        if not pos or pos > ncols: return default
        v = row.iloc[pos - 1]
        if pd.isna(v) or str(v).strip() in ['-','nan','None','']: return default
        return str(v).strip()
    return g

def do_import(uploaded, col_map, skip_rows, sheet):
    suffix = '.xlsx' if uploaded.name.lower().endswith(('.xlsx','.xls')) else '.csv'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded.read())
        tmp_path = tmp.name
    try:
        df = parse_df(tmp_path, sheet=sheet or None)
    finally:
        os.unlink(tmp_path)
    df = df.iloc[skip_rows:].reset_index(drop=True)
    g = read_by_positions(df, col_map)
    amat_idx = col_map['amat_code'] - 1
    if len(df) > 0:
        fv = str(df.iloc[0, amat_idx]).strip()
        if fv and not any(c.isdigit() for c in fv) and len(fv) < 40:
            df = df.iloc[1:].reset_index(drop=True)
    db = get_db()
    cur = db.execute("INSERT INTO imports (filename, col_map) VALUES (?,?)",
                     (uploaded.name, json.dumps(col_map)))
    import_id = cur.lastrowid
    inserted = 0
    for i, row in df.iterrows():
        amat = g(row, 'amat_code')
        if not amat: continue
        try: lv = int(float(g(row,'lv','0')))
        except: lv = 0
        db.execute("""INSERT OR REPLACE INTO parts
            (import_id,row_no,status,amat_parent,amat_code,lv,amat_rev,fv_code,fv_rev,part_name)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (import_id, i+1, g(row,'status'), g(row,'amat_parent'), amat, lv,
             g(row,'amat_rev'), g(row,'fv_code'), g(row,'fv_rev'), g(row,'part_name')))
        inserted += 1
    db.commit()
    load_parts.clear()
    list_imports.clear()
    get_optimized_maps.clear()
    get_parts_lookup.clear()
    return import_id, inserted

@st.cache_data
def load_parts(import_id):
    rows = get_db().execute(
        "SELECT row_no, lv, amat_code, fv_code, fv_rev, amat_rev, part_name, status "
        "FROM parts WHERE import_id=? ORDER BY row_no", (import_id,)
    ).fetchall()
    return [dict(r) for r in rows]

@st.cache_data
def list_imports():
    return [dict(r) for r in get_db().execute(
        "SELECT i.id, i.filename, i.imported_at, COUNT(p.id) as part_count "
        "FROM imports i LEFT JOIN parts p ON p.import_id=i.id "
        "GROUP BY i.id ORDER BY i.id DESC"
    ).fetchall()]

@st.cache_data
def get_optimized_maps(import_id):
    db = get_db()
    rows = db.execute(
        "SELECT amat_code, lv FROM parts WHERE import_id=? ORDER BY row_no",
        (import_id,)
    ).fetchall()
    children_map = defaultdict(list)
    parents_map  = defaultdict(list)
    stack = []
    for row in rows:
        code = str(row["amat_code"]).strip()
        lv   = int(row["lv"] or 0)
        while stack and stack[-1][0] >= lv:
            stack.pop()
        if stack:
            parent = stack[-1][1]
            if code not in children_map[parent]:
                children_map[parent].append(code)
            if parent not in parents_map[code]:
                parents_map[code].append(parent)
        stack.append((lv, code))
    return dict(children_map), dict(parents_map)

@st.cache_data
def get_parts_lookup(import_id):
    rows = get_db().execute(
        "SELECT amat_code, part_name, fv_code, fv_rev, amat_rev, lv, status "
        "FROM parts WHERE import_id=? ORDER BY row_no", (import_id,)
    ).fetchall()
    return {str(r['amat_code']).strip(): dict(r) for r in rows}

def build_tree(node, lookup, visited=None, parts_lookup=None):
    if visited is None: visited = set()
    if node in visited:
        return {"id": node, "children": [], "circular": True}
    visited.add(node)
    p = (parts_lookup or {}).get(node, {})
    return {
        "id":       node,
        "name":     p.get('part_name') or '',
        "fv":       p.get('fv_code') or '',
        "fv_rev":   p.get('fv_rev') or '',
        "amat_rev": p.get('amat_rev') or '',
        "lv":       p.get('lv') or '',
        "status":   p.get('status') or '',
        "children": [build_tree(c, lookup, visited.copy(), parts_lookup) for c in lookup.get(node, [])]
    }

# ── Exports ───────────────────────────────────────────────────────────────────
def export_excel(import_id):
    db = get_db()
    imp = dict(db.execute("SELECT * FROM imports WHERE id=?", (import_id,)).fetchone())
    rows = [dict(r) for r in db.execute(
        "SELECT * FROM parts WHERE import_id=? ORDER BY row_no", (import_id,)).fetchall()]
    wb = Workbook(); ws = wb.active; ws.title = "BOM"
    headers = ['No','Status','AMAT Parent','AMAT Code','LV','AMAT Rev','FV Code','FV Rev','Part Name']
    fields  = ['row_no','status','amat_parent','amat_code','lv','amat_rev','fv_code','fv_rev','part_name']
    lv_colors = {0:'FFF9A825',1:'FFFFF9C4',2:'FFF1F8E9',3:'FFE8F5E9'}
    hf = PatternFill("solid", fgColor="FF1565C0")
    thin = Side(style='thin', color='FFB0BEC5')
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill=hf; cell.font=Font(bold=True,color="FFFFFFFF",size=10)
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
    for w,i in zip([5,10,18,22,5,10,14,10,55], range(1,10)):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in rows:
        lv = r.get('lv',0)
        fill = PatternFill("solid", fgColor=lv_colors.get(lv,'FFFFFFFF'))
        ws.append([r.get(f,'') for f in fields])
        for cell in ws[ws.max_row]:
            cell.fill=fill; cell.border=border
            cell.alignment=Alignment(vertical='center',indent=lv if cell.column==9 else 0)
    ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:I{ws.max_row}"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, imp['filename']+'_export.xlsx'

def export_tree_excel(part, tree_data, direction):
    rows = []
    def walk(node, depth=0):
        rows.append({
            'Level': depth,
            'AMAT Code': node['id'],
            'Part Name': node.get('name',''),
            'FV Code': node.get('fv',''),
            'FV Rev': node.get('fv_rev',''),
            'AMAT Rev': node.get('amat_rev',''),
            'Status': node.get('status',''),
            'Note': '⚠ circular' if node.get('circular') else ''
        })
        for child in node.get('children', []):
            walk(child, depth + 1)
    walk(tree_data)
    wb = Workbook(); ws = wb.active; ws.title = part[:20]
    hf = PatternFill("solid", fgColor="FF1565C0")
    thin = Side(style='thin', color='FFB0BEC5')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    lv_colors = ['FFF9A825','FFFFF9C4','FFF1F8E9','FFE8F5E9','FFEDE7F6']
    headers = ['Level','AMAT Code','Part Name','FV Code','FV Rev','AMAT Rev','Status','Note']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill=hf; cell.font=Font(bold=True,color="FFFFFFFF",size=10)
        cell.alignment=Alignment(horizontal='center'); cell.border=border
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    for r in rows:
        lv = r['Level']
        fill = PatternFill("solid", fgColor=lv_colors[min(lv, 4)])
        ws.append([r['Level'], r['AMAT Code'], r['Part Name'], r['FV Code'],
                   r['FV Rev'], r['AMAT Rev'], r['Status'], r['Note']])
        for cell in ws[ws.max_row]:
            cell.fill=fill; cell.border=border
            cell.alignment=Alignment(vertical='center',
                                     indent=lv if cell.column == 3 else 0)
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def export_tree_text(part, tree_data, direction):
    lines = [f"BOM Tree — {part} — {direction}", '='*60]
    lines.append(part + (f"  {tree_data.get('name','')}" if tree_data.get('name') else ''))
    def walk(node, prefix='', is_last=True):
        connector = '└── ' if is_last else '├── '
        flag = ' ⚠ circular' if node.get('circular') else ''
        extra = ''
        if node.get('fv'): extra += f"  FV:{node['fv']}"
        if node.get('fv_rev'): extra += f" r{node['fv_rev']}"
        if node.get('name'): extra += f"  {node['name']}"
        lines.append(prefix + connector + node['id'] + extra + flag)
        children = node.get('children', [])
        for i, child in enumerate(children):
            walk(child, prefix + ('    ' if is_last else '│   '), i == len(children) - 1)
    for i, child in enumerate(tree_data.get('children', [])):
        walk(child, '', i == len(tree_data.get('children', [])) - 1)
    return '\n'.join(lines)

# ── D3 HTML ───────────────────────────────────────────────────────────────────
def make_tree_html(part, direction, tree_json):
    dir_label = "↓ Explosion" if direction == "explosion" else "↑ Where Used"
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  :root {{--bg:#0d0f14;--surface:#151820;--border:#252a36;--accent:#4fffb0;--accent2:#7c6fff;--text:#e8eaf0;--muted:#5a6070;--warn:#ff6b6b;}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  html,body{{width:100%;height:100%;background:var(--bg);color:var(--text);font-family:monospace,sans-serif;overflow:hidden}}
  #toolbar{{position:absolute;top:0;left:0;right:0;height:46px;background:var(--surface);border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px;padding:0 16px;z-index:10}}
  .root-badge{{background:var(--accent);color:#000;font-size:13px;font-weight:700;padding:3px 10px;border-radius:4px}}
  .dir-badge{{background:var(--accent2);color:#fff;font-size:11px;font-weight:700;padding:3px 9px;border-radius:4px;text-transform:uppercase}}
  .tbtn{{background:var(--border);color:var(--text);border:1px solid var(--border);font-size:11px;font-weight:600;padding:4px 12px;border-radius:4px;cursor:pointer;transition:all .15s}}
  .tbtn:hover{{background:var(--accent);color:#000;border-color:var(--accent)}}
  #stats{{margin-left:auto;font-size:11px;color:var(--muted)}}
  #chart{{position:absolute;top:46px;left:0;right:0;bottom:0;overflow:hidden}}
  #chart svg{{background-image:linear-gradient(var(--border) 1px,transparent 1px),linear-gradient(90deg,var(--border) 1px,transparent 1px);background-size:40px 40px}}
  .node circle{{stroke-width:2px;transition:r .2s,fill .2s;cursor:pointer}}
  .node circle:hover{{r:9}}
  .node.root circle{{fill:var(--accent);stroke:#fff;r:10}}
  .node.leaf circle{{fill:var(--bg);stroke:var(--accent2)}}
  .node.inner circle{{fill:var(--surface);stroke:var(--accent)}}
  .node.collapsed circle{{fill:var(--accent2);stroke:var(--accent2)}}
  .node.circular circle{{fill:var(--warn);stroke:var(--warn)}}
  .node text{{font-family:monospace;font-size:11px;fill:var(--text);pointer-events:none;dominant-baseline:middle}}
  .node.root text{{font-weight:700;font-size:13px;fill:var(--accent)}}
  .link{{fill:none;stroke:var(--border);stroke-width:1.5px}}
  #tooltip{{position:absolute;pointer-events:none;background:var(--surface);border:1px solid var(--accent);border-radius:6px;padding:8px 14px;font-size:12px;color:var(--text);box-shadow:0 4px 20px rgba(0,0,0,.5);display:none;z-index:20;max-width:340px}}
  #tooltip .tp{{font-weight:700;color:var(--accent);font-size:13px;margin-bottom:4px}}
  #tooltip .tn{{color:var(--text);margin-bottom:4px;font-size:11px;white-space:normal;line-height:1.4}}
  #tooltip .tm{{color:var(--muted);margin-top:2px;font-size:10px}}
  #tooltip .tr{{color:var(--accent2);font-size:10px;margin-top:2px}}
</style></head><body>
<div id="toolbar">
  <span class="root-badge">{part}</span>
  <span class="dir-badge">{dir_label}</span>
  <button class="tbtn" id="btn-expand">Expand All</button>
  <button class="tbtn" id="btn-collapse">Collapse All</button>
  <button class="tbtn" id="btn-reset">Reset Zoom</button>
  <span id="stats"></span>
</div>
<div id="chart"></div>
<div id="tooltip"></div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.8.5/d3.min.js"></script>
<script>
const RAW = {tree_json};
const W = window.innerWidth, H = window.innerHeight - 46, DUR = 300;

function toH(n) {{
  return {{
    id: n.id, circular: n.circular||false,
    name: n.name||'', fv: n.fv||'', fv_rev: n.fv_rev||'',
    amat_rev: n.amat_rev||'', lv: n.lv||'', status: n.status||'',
    _children: n.children&&n.children.length ? n.children.map(toH) : null,
    children: null,
  }};
}}

const rootData = toH(RAW);
if(rootData._children){{rootData.children=rootData._children;rootData._children=null;}}
const root = d3.hierarchy(rootData,d=>d.children);
root.x0=H/2;root.y0=0;
const svg=d3.select("#chart").append("svg").attr("width",W).attr("height",H);
const zoom=d3.zoom().scaleExtent([0.1,4]).on("zoom",e=>g.attr("transform",e.transform));
svg.call(zoom);
const g=svg.append("g").attr("transform",`translate(80,${{H/2}})`);
const linkPath=d3.linkHorizontal().x(d=>d.y).y(d=>d.x);
const treeLayout=d3.tree().nodeSize([26,300]);
let nid=0;
function assignIds(d){{if(!d.data._uid)d.data._uid=++nid;if(d.children)d.children.forEach(assignIds);}}
assignIds(root);
const tip=document.getElementById("tooltip");

function update(src){{
  treeLayout(root);
  let total=0,exp=0;root.each(d=>{{total++;if(d.children)exp++;}});
  document.getElementById("stats").textContent=`${{total}} nodes · ${{exp}} expanded`;
  const nodes=root.descendants(),links=root.links();
  const link=g.selectAll(".link").data(links,d=>d.target.data._uid);
  link.enter().insert("path","g").attr("class","link")
    .attr("d",()=>{{const o={{x:src.x0,y:src.y0}};return linkPath({{source:o,target:o}});}})
    .merge(link).transition().duration(DUR).attr("d",linkPath);
  link.exit().transition().duration(DUR).attr("d",()=>{{const o={{x:src.x,y:src.y}};return linkPath({{source:o,target:o}});}}).remove();
  const node=g.selectAll(".node").data(nodes,d=>d.data._uid);
  const ne=node.enter().append("g").attr("class",d=>nc(d))
    .attr("transform",`translate(${{src.y0}},${{src.x0}})`).style("opacity",0)
    .on("click",(e,d)=>{{toggle(d);update(d);}})
    .on("mousemove",(e,d)=>{{
      const k=(d.data._children?.length||0)+(d.data.children?.length||0);
      tip.innerHTML=`
        <div class="tp">${{d.data.id}}</div>
        ${{d.data.name?`<div class="tn">${{d.data.name}}</div>`:''}}
        ${{d.data.fv?`<div class="tm">FV: ${{d.data.fv}}${{d.data.fv_rev?' · rev '+d.data.fv_rev:''}}</div>`:''}}
        ${{d.data.amat_rev?`<div class="tm">AMAT rev: ${{d.data.amat_rev}}</div>`:''}}
        ${{d.data.status?`<div class="tm">Status: ${{d.data.status}}</div>`:''}}
        <div class="tm">L${{d.data.lv}} · depth ${{d.depth}} · ${{k}} children</div>
        <div class="tr">${{d.data.circular?"⚠ circular ref":d.data.children?"click to collapse":d.data._children?"click to expand":"leaf node"}}</div>`;
      tip.style.display="block";tip.style.left=(e.pageX+14)+"px";tip.style.top=(e.pageY-14)+"px";
    }}).on("mouseleave",()=>tip.style.display="none");
  ne.append("circle").attr("r",7);
  ne.append("text").attr("dy","0.31em")
    .attr("x",d=>(d.data.children||d.data._children)?-14:14)
    .attr("text-anchor",d=>(d.data.children||d.data._children)?"end":"start")
    .text(d=>{{
      const name = d.data.name ? '  ' + d.data.name.slice(0,0)+(d.data.name.length>28?'…':'') : '';
      return d.data.id + name;
    }});
  node.merge(ne).transition().duration(DUR).attr("transform",d=>`translate(${{d.y}},${{d.x}})`).style("opacity",1).attr("class",d=>nc(d));
  node.exit().transition().duration(DUR).attr("transform",`translate(${{src.y}},${{src.x}})`).style("opacity",0).remove();
  nodes.forEach(d=>{{d.x0=d.x;d.y0=d.y;}});
}}

function nc(d){{
  if(d.depth===0)return "node root";
  if(d.data.circular)return "node circular";
  if(!d.data._children&&!d.data.children)return "node leaf";
  return d.data.children?"node inner":"node collapsed";
}}

function toggle(d){{
  if(d.data.children){{d.data._children=d.data.children;d.data.children=null;d.children=null;}}
  else if(d.data._children){{d.data.children=d.data._children;d.data._children=null;}}
  const r=d3.hierarchy(rootData,n=>n.children);
  r.each(n=>{{if(!n.x0){{n.x0=H/2;n.y0=0;}}}});Object.assign(root,r);assignIds(root);
}}

document.getElementById("btn-expand").onclick=()=>{{
  root.each(d=>{{if(d.data._children){{d.data.children=d.data._children;d.data._children=null;}}}});
  Object.assign(root,d3.hierarchy(rootData,n=>n.children));root.each(n=>{{if(!n.x0){{n.x0=H/2;n.y0=0;}}}});assignIds(root);update(root);
}};
document.getElementById("btn-collapse").onclick=()=>{{
  root.each(d=>{{if(d.depth>0&&d.data.children){{d.data._children=d.data.children;d.data.children=null;}}}});
  Object.assign(root,d3.hierarchy(rootData,n=>n.children));root.each(n=>{{if(!n.x0){{n.x0=H/2;n.y0=0;}}}});assignIds(root);update(root);
}};
document.getElementById("btn-reset").onclick=()=>svg.transition().duration(400).call(zoom.transform,d3.zoomIdentity.translate(80,H/2));
root.each(d=>{{d.x0=H/2;d.y0=0;}});update(root);
</script></body></html>"""

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""<style>
html,body,[data-testid="stAppViewContainer"]{background:#0d0f14!important}
[data-testid="stSidebar"]{background:#13161b!important;border-right:1px solid #2a2f3a}
[data-testid="stSidebar"] label,[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,[data-testid="stSidebar"] div{color:#e2e8f0!important}
.stTabs [data-baseweb="tab"]{color:#8896aa;font-size:12px}
.stTabs [aria-selected="true"]{color:#4f9cf9!important;border-bottom-color:#4f9cf9!important}
[data-testid="stTextInput"] input{background:#1a1e26!important;border-color:#353c4a!important;color:#e2e8f0!important;font-family:monospace}
.stButton button{background:#1a1e26!important;border:1px solid #353c4a!important;color:#e2e8f0!important;font-size:11px!important}
.stButton button:hover{background:#4f9cf9!important;color:#000!important;border-color:#4f9cf9!important}
[data-testid="stExpander"]{background:#13161b!important;border:1px solid #2a2f3a!important;border-radius:6px!important}
</style>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ BOM/tree")
    st.divider()

    uploaded = st.file_uploader("Upload .xlsx / .csv", type=["xlsx","xls","csv"], label_visibility="collapsed")
    if uploaded:
        sheet_choice = None
        if uploaded.name.lower().endswith(('.xlsx','.xls')):
            uploaded.seek(0)
            sheets = pd.ExcelFile(uploaded).sheet_names
            sheet_choice = st.selectbox("Worksheet", sheets) if len(sheets)>1 else sheets[0]
            uploaded.seek(0)
        with st.expander("Column mapping", expanded=False):
            col_map = {f: st.number_input(FIELD_LABELS[f], min_value=0, max_value=50, value=v, key=f"cm_{f}")
                       for f, v in DEFAULT_COL_MAP.items()}
            skip_rows = st.number_input("Skip rows", min_value=0, max_value=20, value=1)
        if st.button("⬆ Import", use_container_width=True, type="primary"):
            try:
                uploaded.seek(0)
                iid, n = do_import(uploaded, col_map, skip_rows, sheet_choice)
                st.success(f"✓ {n} parts imported")
                st.session_state['active_import'] = iid
                st.rerun()
            except Exception as e:
                st.error(f"Import failed: {e}")

    st.divider()
    st.markdown("**History**")
    imports = list_imports()
    if not imports:
        st.caption("No imports yet")
    else:
        for imp in imports:
            is_active = st.session_state.get('active_import') == imp['id']
            label = f"{'▶ ' if is_active else ''}{imp['filename']}\n{imp['part_count']} parts · {imp['imported_at'][:16]}"
            col_a, col_b = st.columns([5, 1])
            with col_a:
                if st.button(label, key=f"imp_{imp['id']}", use_container_width=True,
                             type="primary" if is_active else "secondary"):
                    st.session_state['active_import'] = imp['id']
                    st.rerun()
            with col_b:
                if st.button("🗑", key=f"del_{imp['id']}", help="Delete"):
                    db = get_db()
                    db.execute("DELETE FROM parts WHERE import_id=?", (imp['id'],))
                    db.execute("DELETE FROM imports WHERE id=?", (imp['id'],))
                    db.commit()
                    load_parts.clear(); list_imports.clear()
                    get_optimized_maps.clear(); get_parts_lookup.clear()
                    if st.session_state.get('active_import') == imp['id']:
                        st.session_state.pop('active_import')
                    st.rerun()

    st.divider()
    active_id = st.session_state.get('active_import')
    if active_id:
        if st.button("↓ Export full BOM", use_container_width=True):
            buf, fname = export_excel(active_id)
            st.download_button("⬇ Click to download", buf, fname,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

# ── Main ──────────────────────────────────────────────────────────────────────
active_id = st.session_state.get('active_import')
if not active_id:
    st.info("Upload and import a BOM file to get started.")
    st.stop()

children_map, parents_map = get_optimized_maps(active_id)

tab_tree, tab_table = st.tabs(["🌳 Tree", "📋 Table"])

with tab_tree:
    col1, col2 = st.columns([3, 1])
    with col1:
        part = st.text_input("Part number", placeholder="e.g. 0011-04827",
                             label_visibility="collapsed", key="search_input").strip().upper()
    with col2:
        direction = st.radio("Direction", ["Explosion ↓", "Where Used ↑"],
                             label_visibility="collapsed", horizontal=True)

    if not part:
        st.info("Enter a part number above to draw its tree.")
    else:
        active_map = children_map if "Explosion" in direction else parents_map
        if part not in children_map and part not in parents_map:
            st.warning(f"Part `{part}` not found.")
        else:
            parts_lookup = get_parts_lookup(active_id)
            tree_data = build_tree(part, active_map, parts_lookup=parts_lookup)
            tree_json = json.dumps(tree_data)
            components.html(
                make_tree_html(part, "explosion" if "Explosion" in direction else "whereused", tree_json),
                height=700, scrolling=False
            )
            ex1, ex2 = st.columns(2)
            with ex1:
                if st.button("⬇ Export tree — Excel", use_container_width=True):
                    buf = export_tree_excel(part, tree_data, direction)
                    st.download_button("⬇ Download Excel", buf, f"{part}_tree.xlsx",
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
            with ex2:
                if st.button("⬇ Export tree — Text", use_container_width=True):
                    txt = export_tree_text(part, tree_data, direction)
                    st.download_button("⬇ Download Text", txt.encode(),
                                       f"{part}_tree.txt", "text/plain",
                                       use_container_width=True)

with tab_table:
    parts = load_parts(active_id)
    df_parts = pd.DataFrame(parts)
    if not df_parts.empty:
        cols = ['row_no','lv','amat_code','fv_code','fv_rev','amat_rev','part_name','status']
        st.dataframe(
            df_parts[[c for c in cols if c in df_parts.columns]],
            use_container_width=True, height=650,
            column_config={
                'lv':        st.column_config.NumberColumn('LV', width='small'),
                'row_no':    st.column_config.NumberColumn('#',  width='small'),
                'amat_code': st.column_config.TextColumn('AMAT Code', width='medium'),
                'fv_code':   st.column_config.TextColumn('FV Code',   width='medium'),
                'part_name': st.column_config.TextColumn('Part Name', width='large'),
            }
        )
